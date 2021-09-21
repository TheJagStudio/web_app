from flask import Blueprint, render_template, request, flash, jsonify,redirect, url_for
from flask_login import login_required, current_user
from .models import Note
from . import db
import json
import openpyxl
import numpy as np

views = Blueprint('views', __name__)
answer_list = [2]
path_questions = "D:/questions.xlsx"
path_marks = "D:/marks.xlsx"
wb_obj = openpyxl.load_workbook(path_questions)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = answer_list[0], column = 1)

webpage = "http://127.0.0.1:5000/"
qarr = [x for x in range(2, sheet_obj.max_row+1)]
np.random.shuffle(qarr)
wb = openpyxl.load_workbook(path_marks)
sheet = wb.active
'''@views.route('/score')
@login_required
def answer(marks):
    wb.save("D:marks.xlsx")
    sheet['A2'].value = current_user.roll_no
    sheet['B2'].value = current_user.first_name
    sheet['C2'].value = marks
    return '<h1>Total marks are :'+marks+'<h1>'''

@views.route('/', methods=['GET', 'POST'])
@login_required
def home():
    print(answer_list[0])
    arr = np.array([2, 3, 4, 5])
    np.random.shuffle(arr)
    print(arr, qarr)
    if request.method == 'POST':
        answer = request.form['q_answer']
        print(chr(68+qarr[answer_list[0]-2]-2)+''+str(current_user.id+1))
        sheet[chr(68+qarr[answer_list[0]-2]-2)+''+str(current_user.id+1)].value = answer
        total_que = sheet_obj.max_row
        if total_que >= answer_list[0]:
            answer_list.append(answer)
            answer_list[0] = answer_list[0] + 1
            print(answer_list)
            if answer_list[0] != (sheet_obj.max_row + 1):
                wb.save("D:marks.xlsx")
                return render_template("home.html", user=current_user,row=answer_list[0], sheet_obj=sheet_obj, arr=arr, qarr=qarr,webpage=webpage)
            else:
                print("went through")
                marks = 0
                temp = sheet_obj.max_row - 1
                row = 2
                while temp != 0:
                    if answer_list[row - 1] in sheet_obj.cell(row=qarr[row - 2], column=6).value:
                        print(marks)
                        marks = marks + 1
                    temp = temp - 1
                    row = row + 1
                sheet['A' + '' + str(current_user.id + 1)].value = current_user.roll_no
                sheet['B' + '' + str(current_user.id + 1)].value = current_user.first_name
                sheet['C' + '' + str(current_user.id + 1)].value = marks
                wb.save("D:/marks.xlsx")
                return '<h1>Total marks are : '+ str(marks) +'<h1>'
        else:
            marks = 0
            temp = sheet_obj.max_row - 1
            row = 2
            while temp != 0:
                if answer_list[row - 1] in sheet_obj.cell(row=row, column=6).value:
                    print(marks)
                    marks = marks + 1
                temp = temp - 1
                row = row + 1
            print(answer_list)
            sheet['A'+''+str(current_user.id+1)].value = current_user.roll_no
            sheet['B'+''+str(current_user.id+1)].value = current_user.first_name
            sheet['C'+''+str(current_user.id+1)].value = marks
            wb.save("D:marks.xlsx")
            return '<h1>Total marks are :' + marks + '<h1>'


    else:
        wb.save("D:marks.xlsx")
        return render_template("home.html", user=current_user,row=answer_list[0], sheet_obj=sheet_obj, arr=arr, qarr=qarr,webpage=webpage)


@views.route('/delete-note', methods=['POST'])
def delete_note():
    note = json.loads(request.data)
    noteId = note['noteId']
    note = Note.query.get(noteId)
    if note:
        if note.user_id == current_user.id:
            db.session.delete(note)
            db.session.commit()

    return jsonify({})
